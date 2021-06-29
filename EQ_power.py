import matplotlib.pyplot as plt
from openpyxl import load_workbook
from plot import *


def mse(array1, array2):
    difference_array = np.subtract(array1, array2)
    squared_array = np.square(difference_array)
    return squared_array.mean()


def add_data_2_new_sheet(dataframe, target_file):
    # create new sheet
    wb = load_workbook(target_file)
    wb.create_sheet("預測電量")
    wb.save(target_file)

    # save data into new sheet
    data_list = [["預測電量"]]
    data_list += [[str(dataframe["預測電量"][i])] for i in range(len(dataframe))]
    wb = load_workbook(target_file)
    sheet = wb.get_sheet_by_name("預測電量")
    for data in data_list:
        sheet.append(data)
    wb.save(target_file)


def load_data():
    raw_data_file = "TY_climate_2017_2018.csv"
    column_name_tt_avg = "TT-Avg(℃)"  # column_name: TT-Avg(℃), MT-Avg(g)
    column_name_mt_avg = "MT-Avg(g)"
    column_name_8ch = "8CH"

    df = pd.read_csv(raw_data_file)

    df_tt = df[column_name_tt_avg].values
    df_tt = df_tt.reshape(-1, 1)

    df_mt = df[column_name_mt_avg].values
    df_mt = df_mt.reshape(-1, 1)

    df_8ch = df[column_name_8ch]
    list_8ch = []
    for i in range(len(df_8ch)):
        try:
            list_8ch.append(int(df_8ch[i]))
        except ValueError:
            list_8ch.append(int(df_8ch[i].replace(",", "")))
    df_8ch = np.array(list_8ch).reshape(-1, 1).astype("float32")

    #  get all data
    all_data_tt, scaler_all_data_tt = data_preprocessing(df_tt)
    all_data_tt_x, _ = create_dataset(all_data_tt)
    all_data_tt_x = all_data_tt_x.reshape(all_data_tt_x.shape[0], 1, 1)

    all_data_mt, scaler_all_data_mt = data_preprocessing(df_mt)
    all_data_mt_x, _ = create_dataset(all_data_mt)
    all_data_mt_x = all_data_mt_x.reshape(all_data_mt_x.shape[0], 1, 1)

    # predict
    model_name_tt = "saved_models_tt_avg/LSTM_002.h5"
    model_name_mt = "saved_models_mt_avg/LSTM_002.h5"
    predictions_tt = predict_func(all_data_tt_x, model_name_tt, scaler_all_data_tt)
    predictions_mt = predict_func(all_data_mt_x, model_name_mt, scaler_all_data_mt)

    # cal enthalpy
    return enthalpy(predictions_tt, predictions_mt), df_8ch[1:]


def power_forecast(file):
    pred_enth, df_8ch = load_data()
    scale = pd.read_excel(file, sheet_name="工作表4")["Unnamed: 33"].values
    scale = scale[1:366].reshape(-1, 1)
    pred_power = pred_enth * scale
    mse_value = mse(pred_power, df_8ch)
    pred_power_list = [float(x) for x in pred_power]
    df_pred_power = pd.DataFrame(
        {
            "預測電量": pred_power_list
        }
    )
    print(f"mse value: {mse_value}")

    # plot
    labels_8ch = {
        "ground_truth": "True 8CH",
        "predict_value": "Predicted 8CH"
    }
    colors_8ch = ["darkviolet", "plum"]
    fig, ax = plt.subplots(figsize=(25, 16))
    fig, ax = plot_all_data(df_8ch, pred_power, fig, ax, colors_8ch, **labels_8ch)
    plt.legend()
    plt.savefig("temp.jpg")
    return df_pred_power


if __name__ == "__main__":
    file_name = "EQ Power 3.xlsx"
    df = power_forecast(file_name)
    add_data_2_new_sheet(df, file_name)
